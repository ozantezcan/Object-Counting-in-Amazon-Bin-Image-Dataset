from __future__ import print_function, division

import torch
from torch.autograd import Variable
import torch.nn as nn
import numpy as np
import matplotlib.pyplot as plt
import time
import copy
import openpyxl

def make_weights_for_balanced_classes(images, nclasses):                        
    count = [0] * nclasses                                                      
    for item in images:                                                         
        count[item[1]] += 1                                                     
    weight_per_class = [0.] * nclasses                                      
    N = float(sum(count))                                                   
    for i in range(nclasses):                                                   
        weight_per_class[i] = N/float(count[i])                            
    weight = [0] * len(images)                                              
    for idx, val in enumerate(images):                                          
        weight[idx] = weight_per_class[val[1]]                                  
    return weight,weight_per_class

def imshow(inp, title=None):
    """Imshow for Tensor."""
    inp = inp.numpy().transpose((1, 2, 0))
    mean = np.array([0.485, 0.456, 0.406])
    std = np.array([0.229, 0.224, 0.225])
    inp = std * inp + mean
    plt.imshow(inp)
    if title is not None:
        plt.title(title)
    plt.pause(0.001)  # pause a bit so that plots are updated

def train_model(model, criterion, optimizer, lr_scheduler,dset_loaders,\
dset_sizes,writer,use_gpu=True, num_epochs=25,batch_size=4,num_log=100,\
init_lr=0.001,lr_decay_epoch=7,multilabel=False,multi_prob=False,mse_loss=False,
cross_loss=1.,multi_loss=0.,
numOut=6, logname='logs.xlsx', iter_loc=12):


    book = openpyxl.load_workbook(logname)
    sheet = book.active
    current_row = sheet.max_row
    since = time.time()

    best_model = model
    best_rmse = 100.0

    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)

        # Each epoch has a training and validation phase
        for phase in ['train', 'val']:
            if phase == 'train':
                batch_count=0
                if lr_scheduler is not None:
                    optimizer = lr_scheduler(optimizer, epoch,init_lr=init_lr,lr_decay_epoch=lr_decay_epoch)
                model.train(True)  # Set model to training mode
            else:
                model.train(False)  # Set model to evaluate mode

            running_loss = 0.0
            running_corrects = 0
            running_cir1=0
            running_mse=0.0
            running_hist=np.zeros(numOut)

            # Iterate over data.
            for data in dset_loaders[phase]:
                # get the inputs
                inputs, labels = data
                if(mse_loss):
                    labels = (labels.type(torch.FloatTensor)-(numOut-1)/2.)/(numOut-1.)*2.

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                        Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                loss = torch.Tensor(1)
                loss = 0.0
                if(mse_loss):
                    #print(labels)
                    tanh_step=torch.nn.Tanh()
                    outputs=tanh_step(outputs)
                    preds=outputs.data
                    #print(preds)
                    criterion = torch.nn.MSELoss()
                    loss += criterion(outputs, labels)
                else:
                    _, preds = torch.max(outputs.data, 1)
                    if cross_loss>0.:
                        criterion=nn.CrossEntropyLoss()
                        loss += cross_loss*criterion(outputs, labels)
                    if multi_loss>0.:
                        labels_multi=[]
                        for label in labels.data:
                            label_multi=np.zeros(numOut+2)


                            if(multi_prob):
                                label_multi[label]=.7
                                label_multi[label+1]=1
                                label_multi[label+2]=.7
                            else:
                                label_multi[label:label+3]=1

                            label_multi=label_multi[1:-1]
                            labels_multi.append(label_multi)

                        labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, numOut)
                        criterion=nn.MultiLabelSoftMarginLoss()
                        loss += multi_loss*criterion(outputs, labelsv)

                # backward + optimize only if in training phase
                if phase == 'train':
                    batch_count+=1
                    if(np.mod(batch_count,num_log)==0):
                        #batch_loss = running_loss / (batch_count*batch_size)
                        batch_acc = running_corrects / (batch_count*batch_size)
                        batch_cir1 = running_cir1 / (batch_count*batch_size)
                        batch_rmse = np.sqrt(running_mse / (batch_count * batch_size))

                        print('{}/{}, acc: {:.4f}, CIR-1: {:.4f}, RMSE: {:.4f}'
                              .format(batch_count,len(dset_loaders['train']),batch_acc,batch_cir1, batch_rmse))

                        '''
                        print(preds.cpu().numpy().transpose())
                        preds_numpy = preds.cpu().numpy() * ((numOut - 1.) / 2.) + ((numOut - 1.) / 2.)
                        preds_numpy = np.round(preds_numpy)
                        labels_numpy = labels.data.cpu().numpy().reshape(-1, 1) * ((numOut - 1.) / 2.) + (
                        (numOut - 1.) / 2.)
                        print(preds_numpy.transpose())
                        print(labels_numpy.transpose())
                        '''

                    loss.backward()
                    optimizer.step()

                # statistics
                running_loss += loss.data[0]

                if(mse_loss):

                    preds_numpy=preds.cpu().numpy()*((numOut-1.)/2.)+((numOut-1.)/2.)
                    labels_numpy=labels.data.cpu().numpy().reshape(-1,1)*((numOut-1.)/2.)+((numOut-1.)/2.)
                    preds_numpy=np.round(preds_numpy)
                    #preds_numpy=np.minimum(np.maximum(1,preds_numpy),7)
                    running_cir1 += np.sum(np.abs(preds_numpy - labels_numpy) <= 1)
                    running_mse += np.sum((preds_numpy - labels_numpy) * (preds_numpy - labels_numpy))
                    running_corrects += np.sum(np.abs(preds_numpy - labels_numpy) < .3)
                else:
                    running_cir1 += torch.sum(torch.abs(preds - labels.data)<=1)
                    running_corrects += torch.sum(preds == labels.data)
                    running_mse += torch.sum((preds - labels.data) * (preds - labels.data))
                for k in range(9):
                    if(torch.sum(labels.data==k)>0):
                        running_hist[k] += torch.sum(torch.abs(preds[labels.data==k] - k)<=1)/torch.sum(labels.data==k)

            epoch_loss = running_loss/dset_sizes[phase]
            epoch_acc = running_corrects / dset_sizes[phase]
            epoch_cir1 = running_cir1 / dset_sizes[phase]
            epoch_rmse = np.sqrt((running_mse / dset_sizes[phase]))
            epoch_hist = running_hist
            writer.add_scalar(phase+' loss',epoch_loss,epoch)
            writer.add_scalar(phase+' accuracy',epoch_acc,epoch)
            writer.add_scalar(phase+' CIR-1',epoch_cir1,epoch)
            writer.add_scalar(phase + 'RMSE', epoch_rmse, epoch)
            writer.add_histogram(phase+' Histogram',epoch_hist,epoch)
            if phase == 'train':
                epoch_loss_tr = epoch_loss
                epoch_acc_tr = epoch_acc
                epoch_rmse_tr = epoch_rmse

            print('{} Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f} RMSE {:.4f}'.format(
                phase, epoch_loss, epoch_acc, epoch_cir1, epoch_rmse))

            # deep copy the model
            if phase == 'val' and epoch_rmse< best_rmse:
                best_rmse = epoch_rmse
                best_model = copy.deepcopy(model)
                sheet.cell(row=current_row, column=iter_loc).value = epoch + 1
                sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
                sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
                sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_acc_tr
                sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_acc
                sheet.cell(row=current_row, column=iter_loc + 5).value = epoch_rmse_tr
                sheet.cell(row=current_row, column=iter_loc + 6).value = epoch_rmse
                book.save(logname)

        print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val RMSE: {:4f}'.format(best_rmse))
    return best_model

def train_model_balanced(model, criterion, optimizer, lr_scheduler,dset_loaders,\
                         dset_sizes,writer,use_gpu=True, num_epochs=25,batch_size=4,\
                         num_train=100, num_test=10,init_lr=0.001,lr_decay_epoch=7,\
                         multilabel=False, multi_prob=False, logname='logs.xlsx', iter_loc=12):
    since = time.time()

    best_model = model
    best_cir1 = 0.0


    book = openpyxl.load_workbook(logname)
    sheet = book.active
    current_row = sheet.max_row
    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)
            # Iterate over data.
        batch_count=0
        #
        if lr_scheduler is not None:
            optimizer = lr_scheduler(optimizer, epoch, init_lr=init_lr, lr_decay_epoch=lr_decay_epoch)
        model.train(True)  # Set model to training mode
        for opt_iter in range(num_test):  
            running_loss = 0.0
            running_corrects = 0
            running_cir1=0

            for k in range(num_train):
                # get the inputs
                inputs, labels = next(iter(dset_loaders['train']))

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                        Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if(multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # backward + optimize only if in training phase           
                loss.backward()
                optimizer.step()

                # statistics
                running_loss += loss.data[0]
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data)<=1)

            epoch_loss = running_loss / (num_train*batch_size)
            epoch_acc = running_corrects / (num_train*batch_size)
            epoch_cir1 = running_cir1 / (num_train*batch_size)
            epoch_loss_tr = epoch_loss
            epoch_cir1_tr = epoch_cir1
            writer.add_scalar('train loss', epoch_loss, epoch)
            writer.add_scalar('train accuracy', epoch_acc, epoch)
            writer.add_scalar('train CIR-1', epoch_cir1, epoch)

            print('{}/{}, Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'
                  .format(opt_iter+1, num_test, epoch_loss, epoch_acc, epoch_cir1))

            # deep copy the model
        model.train(False)  # Set model to evaluate mode
        running_loss = 0.0
        running_corrects = 0
        running_cir1=0
            
        for data in dset_loaders['val']:
            # get the inputs
            inputs, labels = data

            # wrap them in Variable
            if use_gpu:
                inputs, labels = Variable(inputs.cuda()), \
                    Variable(labels.cuda())
            else:
                inputs, labels = Variable(inputs), Variable(labels)

            # zero the parameter gradients
            optimizer.zero_grad()

            # forward
            outputs = model(inputs)
            _, preds = torch.max(outputs.data, 1)
            if (not multilabel):
                loss = criterion(outputs, labels)
            else:
                labels_multi = []
                for label in labels.data:
                    label_multi = np.zeros(11)

                    if (multi_prob):
                        label_multi[label] = .5
                        label_multi[label + 1] = 1
                        label_multi[label + 2] = .5
                    else:
                        label_multi[label:label + 3] = 1

                    label_multi = label_multi[1:-1]
                    labels_multi.append(label_multi)
                labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                loss = criterion(outputs, labelsv)

            # statistics
            running_loss += loss.data[0]
            running_corrects += torch.sum(preds == labels.data)
            running_cir1 += torch.sum(torch.abs(preds - labels.data)<=1)

        epoch_loss = running_loss / dset_sizes['val']
        epoch_acc = running_corrects / dset_sizes['val']
        epoch_cir1 = running_cir1 / dset_sizes['val']
        writer.add_scalar('val loss', epoch_loss, epoch)
        writer.add_scalar('val accuracy', epoch_acc, epoch)
        writer.add_scalar('val CIR-1', epoch_cir1, epoch)

        print('Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))

        if epoch_cir1 > best_cir1:
            best_cir1 = epoch_cir1
            best_model = copy.deepcopy(model)
            sheet.cell(row=current_row, column=iter_loc).value = epoch+1
            sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
            sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
            sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_cir1_tr
            sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_cir1
            book.save(logname)

        print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val Acc: {:4f}'.format(best_cir1))
    del inputs
    del labels
    return best_model


def train_model_both(model, criterion, optimizer, lr_scheduler, dset_loaders_real, \
                         dset_sizes_real, dset_loaders_synthetic, dset_sizes_synthetic,\
                         writer, use_gpu=True, num_epochs=25, batch_size=4, \
                         num_train=100, num_test=10, init_lr=0.001, lr_decay_epoch=7, \
                         multilabel=False, multi_prob=False, logname='logs.xlsx', iter_loc=12):

    book = openpyxl.load_workbook(logname)
    sheet = book.active
    current_row = sheet.max_row

    batch_size = batch_size*2
    since = time.time()

    best_model = model
    best_cir1 = 0.0

    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch, num_epochs - 1))
        print('-' * 10)
        # Iterate over data.
        batch_count = 0
        #
        if lr_scheduler is not None:
            optimizer = lr_scheduler(optimizer, epoch, init_lr=init_lr, lr_decay_epoch=lr_decay_epoch)
        model.train(True)  # Set model to training mode
        for opt_iter in range(num_test):
            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0

            for k in range(num_train):
                # get the inputs
                inputs_real, labels_real = next(iter(dset_loaders_real['train']))
                inputs_synthetic, labels_synthetic = next(iter(dset_loaders_synthetic['train']))
                inputs = torch.cat([inputs_real, inputs_synthetic], 0)
                labels = torch.cat([labels_real, labels_synthetic], 0)

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                                     Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if (multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # backward + optimize only if in training phase
                loss.backward()
                optimizer.step()

                # statistics
                running_loss += loss.data[0]
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)

            epoch_loss = running_loss / (num_train * batch_size)
            epoch_acc = running_corrects / (num_train * batch_size)
            epoch_cir1 = running_cir1 / (num_train * batch_size)
            writer.add_scalar('train loss', epoch_loss, epoch)
            writer.add_scalar('train accuracy', epoch_acc, epoch)
            writer.add_scalar('train CIR-1', epoch_cir1, epoch)

            print('{}/{}, Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'
                  .format(opt_iter + 1, num_test, epoch_loss, epoch_acc, epoch_cir1))

            # deep copy the model
        model.train(False)  # Set model to evaluate mode


        test_count = 0
        for loader in [dset_loaders_real, dset_loaders_synthetic]:
            running_loss = 0.0
            running_corrects = 0
            running_cir1 = 0
            test_count += 1
            for data in loader['val']:
                # get the inputs
                inputs, labels = data

                # wrap them in Variable
                if use_gpu:
                    inputs, labels = Variable(inputs.cuda()), \
                                     Variable(labels.cuda())
                else:
                    inputs, labels = Variable(inputs), Variable(labels)

                # zero the parameter gradients
                optimizer.zero_grad()

                # forward
                outputs = model(inputs)
                _, preds = torch.max(outputs.data, 1)
                if (not multilabel):
                    loss = criterion(outputs, labels)
                else:
                    labels_multi = []
                    for label in labels.data:
                        label_multi = np.zeros(11)

                        if (multi_prob):
                            label_multi[label] = .5
                            label_multi[label + 1] = 1
                            label_multi[label + 2] = .5
                        else:
                            label_multi[label:label + 3] = 1

                        label_multi = label_multi[1:-1]
                        labels_multi.append(label_multi)
                    labelsv = Variable(torch.FloatTensor(labels_multi).cuda()).view(-1, 9)
                    loss = criterion(outputs, labelsv)

                # statistics
                running_loss += loss.data[0]
                running_corrects += torch.sum(preds == labels.data)
                running_cir1 += torch.sum(torch.abs(preds - labels.data) <= 1)
            if(test_count==1):
                epoch_loss = running_loss / dset_sizes_real['val']
                epoch_acc = running_corrects / dset_sizes_real['val']
                epoch_cir1 = running_cir1 / dset_sizes_real['val']
                writer.add_scalar('real val loss', epoch_loss, epoch)
                writer.add_scalar('real val accuracy', epoch_acc, epoch)
                writer.add_scalar('real val CIR-1', epoch_cir1, epoch)
                print('Real Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))
            if (test_count == 2):
                epoch_loss = running_loss / dset_sizes_synthetic['val']
                epoch_acc = running_corrects / dset_sizes_synthetic['val']
                epoch_cir1 = running_cir1 / dset_sizes_synthetic['val']
                writer.add_scalar('synthetic val loss', epoch_loss, epoch)
                writer.add_scalar('synthetic val accuracy', epoch_acc, epoch)
                writer.add_scalar('synthetic val CIR-1', epoch_cir1, epoch)
                print('Synthetic Val Loss: {:.4f} Acc: {:.4f} CIR-1: {:.4f}'.format(epoch_loss, epoch_acc, epoch_cir1))

        if epoch_cir1 > best_cir1:
            best_cir1 = epoch_cir1
            best_model = copy.deepcopy(model)
            sheet.cell(row=current_row, column=iter_loc).value = epoch+1
            #sheet.cell(row=current_row, column=iter_loc + 1).value = epoch_loss_tr
            sheet.cell(row=current_row, column=iter_loc + 2).value = epoch_loss
            #sheet.cell(row=current_row, column=iter_loc + 3).value = epoch_cir1_tr
            sheet.cell(row=current_row, column=iter_loc + 4).value = epoch_cir1
            book.save(logname)

        print()

    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    print('Best val Acc: {:4f}'.format(best_cir1))
    return model


def exp_lr_scheduler(optimizer, epoch, init_lr=0.001, lr_decay_epoch=7):
    """Decay learning rate by a factor of 0.1 every lr_decay_epoch epochs."""
    lr = init_lr * (0.1**(epoch // lr_decay_epoch))

    if epoch % lr_decay_epoch == 0:
        print('LR is set to {}'.format(lr))

    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

    return optimizer

def visualize_model(model, num_images=6):
    images_so_far = 0
    fig = plt.figure()

    for i, data in enumerate(dset_loaders['val']):
        inputs, labels = data
        if use_gpu:
            inputs, labels = Variable(inputs.cuda()), Variable(labels.cuda())
        else:
            inputs, labels = Variable(inputs), Variable(labels)

        outputs = model(inputs)
        _, preds = torch.max(outputs.data, 1)

        for j in range(inputs.size()[0]):
            images_so_far += 1
            ax = plt.subplot(num_images//2, 2, images_so_far)
            ax.axis('off')
            ax.set_title('predicted: {}'.format(dset_classes[labels.data[j]]))
            imshow(inputs.cpu().data[j])

            if images_so_far == num_images:
                return



